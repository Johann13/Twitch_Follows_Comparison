import pprint

import openpyxl

from channel import main_channel_id
from twitch_api import get_channel_follower, get_cred
from twitch_cred import clientID, secret

'''
Gets 800k Follower from a channel
And writes the data into an excel file
This can easily be modified to write the data into another file format
like json or write it directly into a google sheet.
Keep in mind that there are limits to google sheets (and excel)!
'''
if __name__ == '__main__':
    cred = get_cred(client_id=clientID, secret=secret)
    workbook = openpyxl.open('TwitchMCF.xlsx')
    sheet: openpyxl.workbook.workbook.Worksheet = workbook['Tabelle1']
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'name'
    sheet.cell(row=1, column=3).value = 'followed_at'
    workbook.save('TwitchMCF.xlsx')
    page = sheet['D1'].value

    for i in range(100):
        print('start ' + str(i + 1) + '/' + str(100))
        result = get_channel_follower(main_channel_id, cred[0], 800, page, bearer_token=cred[1])
        page = result['page']
        data = result['data']
        sheet['D1'] = page
        v = sheet.max_row
        for j, d in enumerate(data):
            sheet.cell(row=v + j + 1, column=1).value = d['id']
            sheet.cell(row=v + j + 1, column=2).value = d['name']
            sheet.cell(row=v + j + 1, column=3).value = d['followed_at']

        print(str(i + 1) + '/' + str(100) + ' done')
        workbook.save('TwitchMCF.xlsx')
