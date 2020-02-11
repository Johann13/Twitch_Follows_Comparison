import time
from multiprocessing import Lock, Process

import openpyxl
import openpyxl.utils
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from channel import channel_ids, channel_list
from twitch_api import get_channel_follows, get_cred

'''
Gets the follows of users and writes them into an excel file
'''


def fill(fid: int, file_name: str, from_pos: int, user_ids: [str], l: Lock, cred: (str, str)):
    fill_id = str(fid)  # 'fill_' + str(start).rjust(6, '0') + '_' + str(end).rjust(6, '0')
    print(fill_id + ' start')
    s = time.time()
    t = {}
    # workbook = openpyxl.open('Twitch2AB.xlsx')
    # user_id_sheet: Worksheet = workbook[MainChannelFollowerListName]
    time_list = []
    print_percentage = len(user_ids) // 100
    for i, user_id in enumerate(user_ids):
        s_i = time.time()
        t['A' + str(i + from_pos + 1)] = user_id
        result = get_channel_follows(twitch_id=user_id, client_id=cred[0], bearer_token=cred[1])
        data = result['data']
        ids = list(map(lambda x: x['id'], data))
        for j in range(len(channel_ids)):
            yog_id = channel_ids[j]
            if yog_id in ids:
                t[get_column_letter(j + 2) + str(i + from_pos + 1)] = 1
            else:
                t[get_column_letter(j + 2) + str(i + from_pos + 1)] = 0
        time_list.append(round(abs(s_i - time.time()), 2))
        if i % print_percentage == 0:
            p = round((i / len(user_ids)) * 100, 2)
            avg_per_request = abs(round(sum(time_list) / len(time_list), 2))
            avg = round(avg_per_request * (len(user_ids) - i), 2)
            print(fill_id + ' progress: ' + str(p).ljust(4, '0') + '%\t' +
                  'et: ' + str(avg) + 's\t' + 'ts:' + str(round(abs(s - time.time()), 2)).ljust(4, '0') + 's\t' +
                  'apr: ' + str(avg_per_request) + 's')

    print(fill_id + ' queueing to save')
    l.acquire()
    sav = time.time()
    print(fill_id + ' saving')
    workbook2 = openpyxl.open(file_name + 'B.xlsx')
    FollowChannelRelation: Worksheet = workbook2[FollowChannelRelationName]
    for key in t:
        FollowChannelRelation[key] = t[key]
    workbook2.save(file_name + 'B.xlsx')
    # workbook2.close()
    print(fill_id + ' saving took ' + str(round(abs(sav - time.time()), 2)).ljust(4, '0') + 's')
    t.clear()
    print(fill_id + ' took ' + str(round(abs(s - time.time()), 2)).ljust(4, '0') + 's')
    print(fill_id + ' estimated ' + str(round(abs((sum(time_list) / len(time_list)) * len(user_ids)), 2)).ljust(4,
                                                                                                                '0') + 's')
    l.release()
    pass


'''
Splits a given list of twitch user ids into 8 
Starts 8 processes with each sub list
'''


def multi(file_name: str, sheet_name: str, u_ids: [str], offset: int, total_num_of_data: int, cred: [(str, str)]):
    start = time.time()
    wb = openpyxl.Workbook()
    wb.create_sheet(sheet_name)
    wb.save(file_name + '.xlsx')
    wb.save(file_name + 'B.xlsx')
    lock = Lock()
    processes: [Process] = []
    num_of_processes = 8
    data_size = total_num_of_data // num_of_processes
    print('init Processes')
    for i in range(num_of_processes):
        from_pos = (data_size * i)
        print(str(i) + ' from_pos:' + str(from_pos))
        p = Process(target=fill, args=(
            i,
            file_name,
            from_pos,
            u_ids[(data_size * i) + offset:data_size + (data_size * i) + offset],
            lock,
            cred[i % 4]))
        p.start()
        processes.append(p)

    print('init Processes took ' + str(round(abs(start - time.time()), 2)) + ' s')
    for p in processes:
        p.join()
    workbook = openpyxl.open(file_name + 'B.xlsx')
    workbook.save(file_name + '.xlsx')

    print('needed a total ' + str(abs(start - time.time())) + ' s')
    pass


if __name__ == '__main__':
    pass
