import time
from multiprocessing import Lock, Process

import openpyxl.utils
from openpyxl.utils import *
from openpyxl.worksheet.worksheet import Worksheet

from channel import channel_list

'''
Iterates through parts of the excel sheet and adds how many people follow each channel
and how people follow a combination of two channel.
The Result is a 2d Array which is written to an excel file.
'''
def fill(name: str, sheet_name: str, start_from: int, end_at: int, length: int, lock: Lock):
    print(str(start_from) + '|' + str(end_at) + ' start')
    t = {}
    s = time.time()
    workbook = openpyxl.open(name + '.xlsx')
    FollowChannelRelation: Worksheet = workbook[sheet_name]
    for i in range(start_from, end_at):
        print(str(i + 1) + '/' + str(end_at))
        col_letter_a = get_column_letter(i + 2)
        row_pos_a = i + 3
        col_a = FollowChannelRelation[col_letter_a]
        for j in range(i, length):
            col_letter_b = get_column_letter(j + 2)
            row_pos_b = j + 3
            col_b = FollowChannelRelation[col_letter_b]
            v = 0
            for (a, b) in zip(col_a, col_b):
                if a.value is None and b.value is None:
                    break
                else:
                    if a.value == 1 and b.value == 1:
                        v += 1
                pass
            t[get_column_letter(i + 3) + str(row_pos_b)] = v
            t[get_column_letter(j + 3) + str(row_pos_a)] = v

    lock.acquire()
    workbook2 = openpyxl.open(name + 'R.xlsx')
    ChannelRelationTable: Worksheet = workbook2['a']
    for key in t:
        ChannelRelationTable[key] = t[key]
    workbook2.save(name + 'R.xlsx')

    lock.release()
    print(str(start_from) + '|' + str(end_at) + ' took ' + str(abs(s - time.time())) + ' seconds')
    pass

'''
Splits the task into multiple processes
'''
def main(name: str, sheet_name: str, sizes: [(int, int)]):
    start = time.time()
    workbook = openpyxl.Workbook()
    workbook.create_sheet('a')
    workbook.save(name + 'R.xlsx')
    workbook.close()
    print('init took ' + str(abs(start - time.time())) + ' seconds')
    length = len(channel_list)

    lock = Lock()
    processes: [Process] = []
    for size in sizes:
        p = Process(target=fill, args=(name,
                                       sheet_name,
                                       size[0],
                                       size[1],
                                       length,
                                       lock))
        time.sleep(5)
        p.start()
        processes.append(p)

    for p in processes:
        p.join()

    end = time.time()
    print('total took ' + str(abs(start - end)) + ' seconds')


if __name__ == '__main__':
    pass
